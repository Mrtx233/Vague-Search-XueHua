import os
import fasttext
import threading
import shutil
import warnings
import time
from queue import Queue
import logging
import openpyxl
from xlrd import open_workbook  # 仅保留处理XLS格式的依赖

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()

# 忽略不必要的警告
warnings.filterwarnings("ignore", category=UserWarning, message=".*Unable to avoid copy.*")
warnings.filterwarnings("ignore", category=DeprecationWarning, message=".*xlrd.*")

# -------------------- 核心配置 --------------------
MODEL_PATH = "lid.176.bin"  # fasttext语言模型路径
THRESHOLD = 0.5  # 判定某语种的置信度阈值（>该值才判定为该语种）
NUM_WORKERS = 4  # 默认线程数

# 原有特殊目录
ARABIC_DIR = "越南语文件"
OTHER_DIR = "非目标语种文件"
EMPTY_DIR = "无内容文件"
ERROR_DIR = "处理失败文件"

# 支持的文件后缀（移除 xlsb）
SUPPORTED_SUFFIXES = {'xls', 'xlsx', 'xlsm', 'ett', 'et','xlsb'}

# 用户要求的多语种映射（代码 -> 中文名）
LANG_MAP = {
    'de': '德语',
    'en': '英语',
    'es': '西班牙语',
    'fr': '法语',
    'hi': '印地语',
    'id': '印度尼西亚语',
    'it': '意大利语',
    'ja': '日语',
    'ko': '韩语',
    'nl': '荷兰语',
    'pt': '葡萄牙语',
    'ru': '俄语',
    'th': '泰语',
    'vi': '越南语',
    'zh': '中文'
}


# ---------------------------------------------------

def extract_text_from_xls(xls_path):
    """提取XLS格式文件的单元格文本（旧版Excel）"""
    try:
        workbook = open_workbook(xls_path, on_demand=True, data_only=True)
        full_text = []
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            for row_idx in range(sheet.nrows):
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    if cell_value is not None and str(cell_value).strip():
                        full_text.append(str(cell_value).strip())
        workbook.release_resources()
        return "\n".join(full_text)
    except Exception as e:
        logger.error(f"提取XLS文本失败（{xls_path}）: {str(e)}")
        return ""


def extract_text_from_xlsx_xlsm_et_ett(file_path):
    """提取XLSX/XLSM/ET/ETT格式的单元格文本（含WPS格式）"""
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        full_text = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if not hasattr(sheet, 'iter_rows'):
                logger.warning(f"跳过非可读取工作表: {sheet_name}（文件：{os.path.basename(file_path)}）")
                continue
            for row in sheet.iter_rows(values_only=True):
                for cell_value in row:
                    if cell_value is not None and str(cell_value).strip():
                        full_text.append(str(cell_value).strip())
        workbook.close()
        return "\n".join(full_text)
    except Exception as e:
        logger.error(f"提取文本失败（{file_path}）: {str(e)}")
        return ""


def extract_text_by_format(file_path):
    """根据后缀选择对应提取方法"""
    suffix = file_path.split(".")[-1].lower()
    if suffix not in SUPPORTED_SUFFIXES:
        logger.warning(f"不支持的文件格式：{file_path}（仅支持{SUPPORTED_SUFFIXES}）")
        return ""
    if suffix == 'xls':
        return extract_text_from_xls(file_path)
    elif suffix in ['xlsx', 'xlsm', 'et', 'ett']:
        return extract_text_from_xlsx_xlsm_et_ett(file_path)
    else:
        return ""


def detect_language(text, model, lang_map, threshold=THRESHOLD):
    """
    使用 fastText 模型检测语言。
    返回三个值：
      detected_code: 如果检测到为目标语种或阿拉伯语返回对应代码，否则 None
      prob: 如果 detected_code 不为 None，返回对应置信度；否则返回 top1 概率
      top_code: top1 语言代码（始终返回）
    检测逻辑：
      - 先取 top k（k=5）候选，若候选中有 lang_map 中的代码且概率>threshold，直接返回该代码
      - 否则若 top1 是 'ar' 且 top1_prob>threshold，则返回 'ar'（保留阿拉伯语分类）
      - 否则返回 None，并给出 top1 信息
    """
    text = text.strip().replace("\n", " ")
    if not text:
        return None, 0.0, None
    if len(text) > 5000:
        text = text[:5000]

    try:
        k = min(5, max(1, len(lang_map) + 2))
        labels, probs = model.predict(text, k=k)
        labels = [l.replace("__label__", "") for l in labels]
        # 优先匹配目标 LANG_MAP
        for code, prob in zip(labels, probs):
            if code in lang_map and prob > threshold:
                return code, prob, labels[0]
        # 若 top1 为阿拉伯语并且置信度高，也返回 ar（保留原有阿拉伯识别）
        if labels and labels[0] == 'ar' and probs[0] > threshold:
            return 'ar', probs[0], labels[0]
        # 未命中目标语种
        top_code = labels[0] if labels else None
        top_prob = probs[0] if probs else 0.0
        return None, top_prob, top_code
    except Exception as e:
        logger.error(f"语言检测失败: {str(e)}")
        return None, 0.0, None


def move_file_with_retry(src, dst, max_retries=3, delay=1):
    """带重试的文件移动，解决“文件被占用”问题"""
    for i in range(max_retries):
        try:
            shutil.move(src, dst)
            return True
        except PermissionError:
            if i < max_retries - 1:
                logger.warning(f"文件被占用，重试移动「{os.path.basename(src)}」（第{i + 1}次）")
                time.sleep(delay)
            else:
                logger.error(f"多次重试仍无法移动文件：{os.path.basename(src)}")
                return False
        except Exception as e:
            logger.error(f"移动文件失败「{os.path.basename(src)}」: {str(e)}")
            return False


def worker(task_queue, model, output_root):
    """工作线程：处理单个文件的文本提取与语种分类"""
    while True:
        file_path = task_queue.get()
        if file_path is None:
            task_queue.task_done()
            break

        try:
            filename = os.path.basename(file_path)
            logger.info(f"开始处理：{filename}")

            # 1. 提取文本
            text = extract_text_by_format(file_path)

            # 2. 确定分类目录
            if not text:
                target_dir = EMPTY_DIR
                logger.info(f"检测到空内容：{filename} → {target_dir}")
            else:
                detected_code, prob, top_code = detect_language(text, model, LANG_MAP, threshold=THRESHOLD)
                if detected_code:
                    # 若检测到的是我们需要的 LANG_MAP 中的语种
                    if detected_code == 'vi':
                        target_dir = ARABIC_DIR
                        display_name = "越南语"
                    else:
                        display_name = LANG_MAP.get(detected_code, detected_code)
                        target_dir = f"{display_name}"
                    logger.info(f"检测结果：{filename} → {display_name}（代码：{detected_code}，置信度：{prob:.4f}）")
                else:
                    # 未匹配到目标语种且 top1 不是目标（或者置信度低）
                    if top_code:
                        logger.info(f"未匹配目标语种：{filename} top1={top_code} prob={prob:.4f}，归入 {OTHER_DIR}")
                    else:
                        logger.info(f"未能识别语言：{filename} 归入 {OTHER_DIR}")
                    target_dir = OTHER_DIR

            # 3. 移动文件到对应目录（输出根目录下）
            full_target_dir = os.path.join(output_root, target_dir)
            os.makedirs(full_target_dir, exist_ok=True)
            target_path = os.path.join(full_target_dir, filename)

            # 处理文件名冲突
            counter = 1
            name, ext = os.path.splitext(filename)
            while os.path.exists(target_path):
                target_path = os.path.join(full_target_dir, f"{name}_{counter}{ext}")
                counter += 1

            # 带重试移动文件
            if not move_file_with_retry(file_path, target_path):
                err_dir = os.path.join(output_root, ERROR_DIR)
                os.makedirs(err_dir, exist_ok=True)
                err_path = os.path.join(err_dir, filename)
                move_file_with_retry(file_path, err_path)

            logger.info(f"处理完成：{filename} → {full_target_dir}")

        except Exception as e:
            logger.error(f"处理异常（{file_path}）: {str(e)}")
            error_dir = os.path.join(output_root, ERROR_DIR)
            os.makedirs(error_dir, exist_ok=True)
            error_path = os.path.join(error_dir, os.path.basename(file_path))
            move_file_with_retry(file_path, error_path)
        finally:
            task_queue.task_done()


def batch_process_files(source_dir, output_root, num_workers=NUM_WORKERS):
    """批量处理源目录中所有支持的文件格式"""
    if not os.path.exists(MODEL_PATH):
        logger.error(f"语言模型不存在！请将「lid.176.bin」放在脚本同目录或修改 MODEL_PATH")
        return

    try:
        logger.info("加载语言检测模型...")
        model = fasttext.load_model(MODEL_PATH)
        logger.info("模型加载成功")
    except Exception as e:
        logger.error(f"模型加载失败: {str(e)}")
        return

    # 扫描支持文件
    supported_files = []
    for filename in os.listdir(source_dir):
        file_path = os.path.join(source_dir, filename)
        if os.path.isfile(file_path):
            suffix = filename.split(".")[-1].lower()
            if suffix in SUPPORTED_SUFFIXES:
                supported_files.append(file_path)

    if not supported_files:
        logger.warning(f"源目录「{source_dir}」中未找到支持的文件（仅支持{SUPPORTED_SUFFIXES}）")
        return
    logger.info(f"共发现 {len(supported_files)} 个支持的文件待处理")

    # 初始化线程与任务队列
    task_queue = Queue()
    threads = []
    for i in range(num_workers):
        t = threading.Thread(target=worker, args=(task_queue, model, output_root), name=f"Worker-{i + 1}")
        t.daemon = True
        t.start()
        threads.append(t)
        logger.info(f"启动线程：{t.name}")

    # 加入任务
    for file_path in supported_files:
        task_queue.put(file_path)

    # 等待完成
    task_queue.join()
    logger.info("所有文件处理完成")

    # 发送退出信号并关闭线程
    for _ in range(num_workers):
        task_queue.put(None)
    for t in threads:
        t.join()
    logger.info("所有线程已关闭")


if __name__ == "__main__":
    # 请根据实际情况修改下面两项路径
    SOURCE_DIRECTORY = r"D:\数据采集\0213\常规\ET组件"  # 源文件目录
    OUTPUT_ROOT = SOURCE_DIRECTORY  # 结果保存目录（也可以改为其它路径）
    NUM_WORKERS = 4

    print(f"===== 多语种Excel分类工具 =====")
    print(f"识别目标语种：{', '.join([f'{v}({k})' for k, v in LANG_MAP.items()])}")
    print(f"支持格式：{SUPPORTED_SUFFIXES}")
    print(f"源目录：{SOURCE_DIRECTORY}")
    print(f"结果目录：{OUTPUT_ROOT}")
    print(f"线程数：{NUM_WORKERS}")
    print("======================================")

    batch_process_files(SOURCE_DIRECTORY, OUTPUT_ROOT, NUM_WORKERS)

    print("\n===== 处理完成！=====")
    print(f"结果目录：{OUTPUT_ROOT}")
    print(f"1. 每个目标语种会生成对应文件夹，例如：德语 (de)")
    print(f"2. {ARABIC_DIR}：检测为阿拉伯语的文件")
    print(f"3. {OTHER_DIR}：未命中目标列表或置信度不足的文件")
    print(f"4. {EMPTY_DIR}：未提取到有效文本的文件")
    print(f"5. {ERROR_DIR}：无法正常处理的文件")
